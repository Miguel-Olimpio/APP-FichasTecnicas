"""Migração legado → IngredientesFicha + banco_ingredientes."""

from __future__ import annotations

import os
import tempfile
import uuid

import pytest
from openpyxl import Workbook, load_workbook

from app.config.settings import (
    ORIGEM_CADASTRO_MESTRE,
    ORIGEM_FICHA_INTERMEDIARIA,
    SHEET_INGREDIENTES_FICHA,
    SHEET_INGREDIENTES_LEGACY,
    SHEET_INGREDIENTES_MESTRE,
    SHEET_PRODUTOS,
)
from app.repositories.excel_database import ExcelDatabase, _read_sheet_dicts
from app.repositories.excel_schema import (
    FICHAS_SHEETS_CONFIG,
    LEGACY_INGREDIENT_HEADERS,
    MASTER_SHEETS_CONFIG,
    PRODUCT_HEADERS,
    RECIPE_LINE_HEADERS,
)
from app.repositories.migration_split_workbooks import _needs_split_migration, run_split_workbook_migration
from app.utils.tipo_ficha import TIPO_FICHA_PRODUTO_FINAL


def _write_legacy_fichas(path: str, produto_id: str, ref_id: str) -> tuple[str, str]:
    """Workbook só com Produtos + aba legada Ingredientes (simples + composto)."""
    wb = Workbook()
    wb.remove(wb.active)
    ws_p = wb.create_sheet(SHEET_PRODUTOS)
    ws_p.append(PRODUCT_HEADERS)
    prod_row = {h: "" for h in PRODUCT_HEADERS}
    prod_row.update(
        {
            "produto_id": produto_id,
            "nome": "Produto teste",
            "nome_normalizado": "",
            "categoria": "Outros",
            "tipo_ficha": TIPO_FICHA_PRODUTO_FINAL,
            "rendimento": 1,
            "unidade_rendimento": "kg",
            "quantidade_porcoes": 1,
            "tempo_preparo": "",
            "temperatura": "",
            "modo_preparo": "",
            "observacoes": "",
            "custo_total": 0,
            "custo_por_unidade": 0,
            "custo_por_porcao": 0,
            "custo_por_kg": 0,
            "data_criacao": "2020-01-01",
            "data_atualizacao": "2020-01-01",
            "active": True,
        }
    )
    ws_p.append([prod_row[h] for h in PRODUCT_HEADERS])

    ws_i = wb.create_sheet(SHEET_INGREDIENTES_LEGACY)
    ws_i.append(LEGACY_INGREDIENT_HEADERS)
    ing_simple = str(uuid.uuid4())
    ing_comp = str(uuid.uuid4())
    ws_i.append(
        [
            ing_simple,
            produto_id,
            "Farinha",
            "",
            "simples",
            "",
            1,
            "kg",
            0,
            10,
            "kg",
            0,
            0,
            "",
        ]
    )
    ws_i.append(
        [
            ing_comp,
            produto_id,
            "Base interna",
            "",
            "produto_composto",
            ref_id,
            0.5,
            "kg",
            0,
            0,
            "kg",
            0,
            0,
            "",
        ]
    )
    wb.save(path)
    return ing_simple, ing_comp


@pytest.fixture
def split_env():
    with tempfile.TemporaryDirectory() as tmp:
        fichas = os.path.join(tmp, "banco_fichas.xlsx")
        master = os.path.join(tmp, "banco_ingredientes.xlsx")
        pid = str(uuid.uuid4())
        ref = str(uuid.uuid4())
        _write_legacy_fichas(fichas, pid, ref)

        db_m = ExcelDatabase(master, MASTER_SHEETS_CONFIG, "banco_ingredientes")
        db_m.create_database()

        db_f = ExcelDatabase(fichas, FICHAS_SHEETS_CONFIG, "banco_fichas")
        yield db_f, db_m, fichas, master, pid, ref


def test_needs_split_when_legacy_ingredientes_present(split_env):
    db_f, *_rest = split_env
    assert _needs_split_migration(db_f) is True


def test_migration_simple_to_master_and_composto_produto_ref(split_env):
    db_f, db_m, fichas_path, master_path, produto_id, ref_id = split_env
    warns = run_split_workbook_migration(db_f, db_m)
    assert warns

    wb_f = load_workbook(fichas_path, read_only=True, data_only=True)
    try:
        assert SHEET_INGREDIENTES_LEGACY not in wb_f.sheetnames
        assert SHEET_INGREDIENTES_FICHA in wb_f.sheetnames
        lines = _read_sheet_dicts(wb_f[SHEET_INGREDIENTES_FICHA])
        assert len(lines) == 2
        by_origem = {str(r.get("origem_linha", "")): r for r in lines}
        assert ORIGEM_CADASTRO_MESTRE in by_origem
        assert ORIGEM_FICHA_INTERMEDIARIA in by_origem
        simple = by_origem[ORIGEM_CADASTRO_MESTRE]
        assert str(simple.get("nome_ingrediente", "")).lower() == "farinha"
        assert str(simple.get("ingrediente_id", ""))  # FK mestre
        comp = by_origem[ORIGEM_FICHA_INTERMEDIARIA]
        assert str(comp.get("produto_ref_id", "")) == ref_id
        assert float(simple.get("preco_kg_snapshot", 0) or 0) == 10.0
    finally:
        wb_f.close()

    wb_m = load_workbook(master_path, read_only=True, data_only=True)
    try:
        assert SHEET_INGREDIENTES_MESTRE in wb_m.sheetnames
        masters = _read_sheet_dicts(wb_m[SHEET_INGREDIENTES_MESTRE])
        nomes = {str(m.get("nome", "")).lower() for m in masters}
        assert "farinha" in nomes
    finally:
        wb_m.close()


def test_migration_idempotent_second_run(split_env):
    db_f, db_m, *_ = split_env
    run_split_workbook_migration(db_f, db_m)
    assert _needs_split_migration(db_f) is False
    warns2 = run_split_workbook_migration(db_f, db_m)
    assert warns2 == []
