"""Acesso ao workbook Excel: criação, migração, backup e gravação atômica."""

from __future__ import annotations

import os
import shutil
import uuid
from datetime import datetime
from datetime import timedelta
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from app.config.settings import (
    MSG_SAVE_EXCEL_LOCKED,
    SHEET_INGREDIENTES_FICHA,
    SHEET_INGREDIENTES_MESTRE,
    SHEET_PRODUTOS,
)
from app.repositories.excel_schema import (
    FICHAS_SHEETS_CONFIG,
    MASTER_INGREDIENT_HEADERS,
    PRODUCT_HEADERS,
    RECIPE_LINE_HEADERS,
)
from app.utils.filenames import normalize_name_key
from app.utils.tipo_ficha import TIPO_FICHA_PRODUTO_FINAL


MAX_BACKUPS_PER_DATABASE = 10
MIN_BACKUP_INTERVAL = timedelta(minutes=10)
BACKUP_TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"


class ExcelSaveError(Exception):
    """Erro ao persistir o Excel (ex.: arquivo aberto no Excel)."""

    def __init__(self, message: str = MSG_SAVE_EXCEL_LOCKED):
        super().__init__(message)
        self.user_message = message


def _backup_filename(db_path: str, stem: str = "banco", backup_dir: str | None = None) -> str:
    folder = backup_dir or os.path.dirname(db_path)
    os.makedirs(folder or ".", exist_ok=True)
    stamp = datetime.now().strftime(BACKUP_TIMESTAMP_FORMAT)
    return os.path.join(folder, f"{stem}_backup_{stamp}.xlsx")


def _backup_timestamp(path: str, stem: str) -> datetime | None:
    name = os.path.basename(path)
    prefix = f"{stem}_backup_"
    suffix = ".xlsx"
    if not (name.startswith(prefix) and name.endswith(suffix)):
        return None
    raw = name[len(prefix) : -len(suffix)]
    try:
        return datetime.strptime(raw, BACKUP_TIMESTAMP_FORMAT)
    except ValueError:
        return None


def _list_backups(db_path: str, stem: str, backup_dir: str | None = None) -> list[tuple[datetime, str]]:
    folder = backup_dir or os.path.dirname(db_path) or "."
    if not os.path.isdir(folder):
        return []
    backups: list[tuple[datetime, str]] = []
    for name in os.listdir(folder):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        stamp = _backup_timestamp(path, stem)
        if stamp is not None:
            backups.append((stamp, path))
    backups.sort(key=lambda item: item[0])
    return backups


def _cleanup_old_backups(
    db_path: str,
    stem: str,
    max_backups: int = MAX_BACKUPS_PER_DATABASE,
    backup_dir: str | None = None,
) -> None:
    backups = _list_backups(db_path, stem, backup_dir)
    excess = len(backups) - max_backups
    if excess <= 0:
        return
    for _stamp, path in backups[:excess]:
        try:
            os.remove(path)
        except OSError:
            pass


def _read_sheet_dicts(ws: Worksheet) -> list[dict[str, Any]]:
    if ws.max_row < 1:
        return []
    headers: list[str | None] = []
    for cell in ws[1]:
        headers.append(cell.value)
    while headers and headers[-1] is None:
        headers.pop()
    headers_str = [str(h).strip().lower() for h in headers if h is not None]
    if not headers_str:
        return []
    data: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, max_col=len(headers_str), values_only=True):
        if row is None or all(v is None for v in row):
            continue
        data.append(dict(zip(headers_str, row)))
    return data


def _write_sheet(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for item in rows:
        ws.append([item.get(h, "") for h in headers])


def _merge_row(canonical: list[str], row: dict[str, Any], defaults: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {k: defaults.get(k, "") for k in canonical}
    for k, v in row.items():
        if k in out and v not in (None, ""):
            out[k] = v
    for k in canonical:
        if out[k] in (None, "") and k in defaults:
            out[k] = defaults[k]
    return out


def _migrate_products(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for r in rows:
        nome = r.get("nome", "")
        nn = r.get("nome_normalizado") or normalize_name_key(nome)
        qp = r.get("quantidade_porcoes")
        if qp in (None, ""):
            qp = 1
        active = r.get("active")
        if active in (None, ""):
            active = True
        cpp = r.get("custo_por_porcao")
        if cpp in (None, ""):
            cpp = 0
        tf = r.get("tipo_ficha")
        if tf in (None, ""):
            tf = TIPO_FICHA_PRODUTO_FINAL
        defaults: dict[str, Any] = {
            "nome_normalizado": nn,
            "quantidade_porcoes": qp,
            "active": active,
            "custo_por_porcao": cpp,
            "tipo_ficha": tf,
        }
        result.append(_merge_row(PRODUCT_HEADERS, r, defaults))
    return result


def _migrate_recipe_lines(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for r in rows:
        obs = r.get("observacoes", "")
        if obs is None:
            obs = ""
        defaults: dict[str, Any] = {"observacoes": obs}
        result.append(_merge_row(RECIPE_LINE_HEADERS, r, defaults))
    return result


def _migrate_master_ingredients(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for r in rows:
        nome = r.get("nome", "")
        nn = r.get("nome_normalizado") or normalize_name_key(nome)
        active = r.get("active")
        if active in (None, ""):
            active = True
        pl = r.get("preco_litro")
        if pl in (None, ""):
            pl = 0
        defaults: dict[str, Any] = {"nome_normalizado": nn, "active": active, "preco_litro": pl}
        result.append(_merge_row(MASTER_INGREDIENT_HEADERS, r, defaults))
    return result


class ExcelDatabase:
    def __init__(
        self,
        db_path: str,
        sheets_config: dict[str, list[str]] | None = None,
        backup_stem: str = "banco",
        backup_dir: str | None = None,
    ):
        self._db_path = db_path
        self._sheets_config = sheets_config or FICHAS_SHEETS_CONFIG
        self._backup_stem = backup_stem
        self._backup_dir = backup_dir

    @property
    def db_path(self) -> str:
        return self._db_path

    @property
    def sheets_config(self) -> dict[str, list[str]]:
        return self._sheets_config

    def create_backup(self) -> str | None:
        if not os.path.isfile(self._db_path):
            return None
        backups = _list_backups(self._db_path, self._backup_stem, self._backup_dir)
        now = datetime.now()
        if backups and now - backups[-1][0] < MIN_BACKUP_INTERVAL:
            _cleanup_old_backups(self._db_path, self._backup_stem, backup_dir=self._backup_dir)
            return None
        dest = _backup_filename(self._db_path, self._backup_stem, self._backup_dir)
        shutil.copy2(self._db_path, dest)
        _cleanup_old_backups(self._db_path, self._backup_stem, backup_dir=self._backup_dir)
        return dest

    def create_database(self) -> None:
        os.makedirs(os.path.dirname(self._db_path) or ".", exist_ok=True)
        wb = Workbook()
        first = True
        for name, headers in self._sheets_config.items():
            if first:
                ws = wb.active
                ws.title = name
                ws.append(headers)
                first = False
            else:
                wb.create_sheet(name).append(headers)
        self.save_workbook_safe(wb)

    def load_workbook_safe(self) -> WorkbookType:
        return load_workbook(self._db_path, read_only=False, data_only=False)

    def save_workbook_safe(self, wb: WorkbookType) -> None:
        folder = os.path.dirname(self._db_path) or "."
        os.makedirs(folder, exist_ok=True)
        tmp = f"{self._db_path}.{uuid.uuid4().hex}.tmp"
        try:
            wb.save(tmp)
            wb.close()
            os.replace(tmp, self._db_path)
        except (PermissionError, OSError) as exc:
            if os.path.isfile(tmp):
                try:
                    os.remove(tmp)
                except OSError:
                    pass
            raise ExcelSaveError() from exc

    def read_sheet(self, sheet_name: str) -> list[dict[str, Any]]:
        wb = self.load_workbook_safe()
        try:
            if sheet_name not in wb.sheetnames:
                return []
            return _read_sheet_dicts(wb[sheet_name])
        finally:
            wb.close()

    def write_sheet(self, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
        wb = self.load_workbook_safe()
        try:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            _write_sheet(ws, headers, rows)
            self.save_workbook_safe(wb)
            return
        except ExcelSaveError:
            wb.close()
            raise
        except Exception:
            wb.close()
            raise

    def sheet_names(self) -> list[str]:
        wb = self.load_workbook_safe()
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def remove_sheet(self, sheet_name: str) -> None:
        wb = self.load_workbook_safe()
        try:
            if sheet_name in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb[sheet_name]
            self.save_workbook_safe(wb)
        except ExcelSaveError:
            wb.close()
            raise
        except Exception:
            wb.close()
            raise

    def append_audit_row(self, row: dict[str, Any]) -> None:
        from app.config.settings import SHEET_AUDIT
        from app.repositories.excel_schema import AUDIT_HEADERS

        rows = self.read_sheet(SHEET_AUDIT)
        rows.append(row)
        self.write_sheet(SHEET_AUDIT, AUDIT_HEADERS, rows)

    def migrate_database(self) -> None:
        self.create_backup()
        wb = self.load_workbook_safe()
        try:
            for name, headers in self._sheets_config.items():
                if name not in wb.sheetnames:
                    ws = wb.create_sheet(name)
                    ws.append(headers)
                    continue
                ws = wb[name]
                raw = _read_sheet_dicts(ws)
                if name == SHEET_PRODUTOS:
                    normalized = _migrate_products(raw)
                elif name == SHEET_INGREDIENTES_FICHA:
                    normalized = _migrate_recipe_lines(raw)
                elif name == SHEET_INGREDIENTES_MESTRE:
                    normalized = _migrate_master_ingredients(raw)
                else:
                    normalized = [_merge_row(headers, r, {}) for r in raw]
                _write_sheet(ws, headers, normalized)
            self.save_workbook_safe(wb)
        except ExcelSaveError:
            wb.close()
            raise
        except Exception:
            wb.close()
            raise

    def ensure_database(self) -> None:
        if not os.path.isfile(self._db_path):
            self.create_database()
            return
        self.migrate_database()
