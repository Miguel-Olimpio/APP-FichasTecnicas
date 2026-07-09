"""Migração one-shot: banco_fichas legado (Ingredientes + Catalogo) -> v2 (IngredientesFicha + banco_ingredientes)."""

from __future__ import annotations

import uuid
from typing import Any

from openpyxl import load_workbook

from app.config.settings import (
    CLASSIFICACAO_INGREDIENTE_SIMPLES,
    CLASSIFICACAO_MATERIA_PRIMA,
    ORIGEM_CADASTRO_MESTRE,
    ORIGEM_FICHA_INTERMEDIARIA,
    SHEET_CATALOGO_LEGACY,
    SHEET_INGREDIENTES_FICHA,
    SHEET_INGREDIENTES_LEGACY,
    SHEET_INGREDIENTES_MESTRE,
    SHEET_PRODUTOS,
)
from app.models.enums import IngredienteTipo
from app.repositories.excel_database import ExcelDatabase, _read_sheet_dicts, _write_sheet
from app.repositories.excel_schema import (
    LEGACY_CATALOG_HEADERS,
    MASTER_INGREDIENT_HEADERS,
    PRODUCT_HEADERS,
    RECIPE_LINE_HEADERS,
)
from app.services.recipe_line_adapter import service_dict_to_row
from app.utils.dates import now_str
from app.utils.filenames import normalize_name_key
from app.utils.numbers import to_float
from app.utils.tipo_ficha import TIPO_FICHA_MATERIA_PRIMA, normalize_tipo_ficha


def _normalize_cost_uc(uc: str) -> str:
    s = str(uc or "").strip().lower()
    if s in ("l", "litro", "litros"):
        return "l"
    if s in ("ml", "mililitro", "mililitros"):
        return "ml"
    if s in ("un", "unidade", "unidades"):
        return "un"
    if s in ("porcao", "porção", "porcoes", "porções"):
        return "porção"
    return "kg"


def _needs_split_migration(fichas_db: ExcelDatabase) -> bool:
    wb = load_workbook(fichas_db.db_path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if SHEET_INGREDIENTES_LEGACY in names:
            return True
        if SHEET_INGREDIENTES_FICHA in names:
            rows = _read_sheet_dicts(wb[SHEET_INGREDIENTES_FICHA])
            if rows and "origem_linha" not in rows[0]:
                return True
        return False
    finally:
        wb.close()


def run_split_workbook_migration(fichas_db: ExcelDatabase, master_db: ExcelDatabase) -> list[str]:
    warnings: list[str] = []
    if not _needs_split_migration(fichas_db):
        return warnings

    fichas_db.create_backup()
    master_db.create_backup()

    wb = load_workbook(fichas_db.db_path, read_only=False, data_only=False)
    try:
        legacy_rows: list[dict[str, Any]] = []
        if SHEET_INGREDIENTES_LEGACY in wb.sheetnames:
            legacy_rows = _read_sheet_dicts(wb[SHEET_INGREDIENTES_LEGACY])

        catalog_rows: list[dict[str, Any]] = []
        if SHEET_CATALOGO_LEGACY in wb.sheetnames:
            catalog_rows = _read_sheet_dicts(wb[SHEET_CATALOGO_LEGACY])

        master_rows = master_db.read_sheet(SHEET_INGREDIENTES_MESTRE)
        nn_to_master_id: dict[str, str] = {}
        for r in master_rows:
            nn = str(r.get("nome_normalizado", "") or "").strip().lower()
            if nn:
                nn_to_master_id[nn] = str(r.get("ingrediente_id", "") or "")

        def upsert_master_from_simple(nome: str, uc: str, puk: float, puu: float, classificacao: str) -> str:
            nn = normalize_name_key(nome)
            key = nn.strip().lower()
            if key in nn_to_master_id:
                return nn_to_master_id[key]
            mid = str(uuid.uuid4())
            ck = _normalize_cost_uc(uc)
            upad = "kg"
            p_litro = 0.0
            p_kg = puk
            p_un = puu
            if ck == "un":
                upad = "un"
                p_kg = 0.0
            elif ck in ("l", "ml"):
                upad = "L"
                p_kg = 0.0
                p_un = 0.0
                p_litro = puk if ck == "l" else 0.0
            row = {
                "ingrediente_id": mid,
                "nome": nome.strip(),
                "nome_normalizado": nn,
                "classificacao": classificacao,
                "categoria": "Outros",
                "unidade_padrao": upad,
                "unidade_custo": uc or "kg",
                "preco_kg": p_kg,
                "preco_litro": p_litro,
                "preco_unidade": p_un,
                "observacoes": "",
                "data_criacao": now_str(),
                "data_atualizacao": now_str(),
                "active": True,
            }
            master_rows.append(row)
            nn_to_master_id[key] = mid
            return mid

        new_recipe: list[dict[str, Any]] = []

        for r in legacy_rows:
            tipo = str(r.get("tipo", IngredienteTipo.SIMPLES.value) or "")
            pid = str(r.get("produto_id", "") or "")
            nome = str(r.get("nome", "") or "")
            q = to_float(r.get("quantidade"))
            u = str(r.get("unidade", "") or "")
            uc = str(r.get("unidade_custo", "") or "")
            puk = to_float(r.get("preco_kg"))
            puu = to_float(r.get("preco_unidade"))
            line_id = str(r.get("ingrediente_id") or uuid.uuid4())
            if tipo == IngredienteTipo.PRODUTO_COMPOSTO.value:
                ref = str(r.get("produto_ref_id", "") or "")
                d = {
                    "ingrediente_ficha_id": line_id,
                    "produto_id": pid,
                    "origem_linha": ORIGEM_FICHA_INTERMEDIARIA,
                    "ingrediente_cadastro_id": "",
                    "produto_ref_id": ref,
                    "nome": nome,
                    "classificacao_ingrediente": "produto_intermediario",
                    "tipo": IngredienteTipo.PRODUTO_COMPOSTO.value,
                    "quantidade": q,
                    "unidade": u,
                    "unidade_custo": uc,
                    "preco_kg": 0.0,
                    "preco_unidade": 0.0,
                    "preco_litro": 0.0,
                    "custo_calculado": to_float(r.get("custo_calculado")),
                    "proporcao": to_float(r.get("proporcao")),
                    "observacoes": str(r.get("observacoes", "") or ""),
                }
                new_recipe.append(service_dict_to_row(d))
            else:
                mid = upsert_master_from_simple(
                    nome, uc, puk, puu, CLASSIFICACAO_INGREDIENTE_SIMPLES
                )
                d = {
                    "ingrediente_ficha_id": line_id,
                    "produto_id": pid,
                    "origem_linha": ORIGEM_CADASTRO_MESTRE,
                    "ingrediente_cadastro_id": mid,
                    "produto_ref_id": "",
                    "nome": nome,
                    "classificacao_ingrediente": CLASSIFICACAO_INGREDIENTE_SIMPLES,
                    "tipo": IngredienteTipo.SIMPLES.value,
                    "quantidade": q,
                    "unidade": u,
                    "unidade_custo": uc,
                    "preco_kg": puk,
                    "preco_unidade": puu,
                    "preco_litro": 0.0,
                    "custo_calculado": to_float(r.get("custo_calculado")),
                    "proporcao": to_float(r.get("proporcao")),
                    "observacoes": str(r.get("observacoes", "") or ""),
                }
                new_recipe.append(service_dict_to_row(d))

        for c in catalog_rows:
            nome = str(c.get("nome", "") or "")
            if not nome.strip():
                continue
            uc = str(c.get("unidade_custo_padrao", "kg") or "kg")
            puk = to_float(c.get("preco_kg_padrao"))
            puu = to_float(c.get("preco_unidade_padrao"))
            upsert_master_from_simple(nome, uc, puk, puu, CLASSIFICACAO_INGREDIENTE_SIMPLES)

        products = _read_sheet_dicts(wb[SHEET_PRODUTOS]) if SHEET_PRODUTOS in wb.sheetnames else []
        for p in products:
            if normalize_tipo_ficha(p.get("tipo_ficha")) != TIPO_FICHA_MATERIA_PRIMA:
                continue
            nome = str(p.get("nome", "") or "")
            if not nome.strip():
                continue
            nn = normalize_name_key(nome).strip().lower()
            if nn in nn_to_master_id:
                continue
            mid = str(uuid.uuid4())
            master_rows.append(
                {
                    "ingrediente_id": mid,
                    "nome": nome.strip(),
                    "nome_normalizado": normalize_name_key(nome),
                    "classificacao": CLASSIFICACAO_MATERIA_PRIMA,
                    "categoria": str(p.get("categoria", "") or "Outros"),
                    "unidade_padrao": "kg",
                    "unidade_custo": "kg",
                    "preco_kg": to_float(p.get("custo_por_kg")),
                    "preco_litro": 0.0,
                    "preco_unidade": to_float(p.get("custo_por_unidade")),
                    "observacoes": "Migrado automaticamente a partir da ficha (matéria-prima).",
                    "data_criacao": now_str(),
                    "data_atualizacao": now_str(),
                    "active": True,
                }
            )
            nn_to_master_id[nn] = mid

        warnings.append(
            "Migração concluída: ingredientes simples foram para o cadastro mestre; "
            "revise classificações e preços em «Ingredientes cadastrados»."
        )

        master_db.write_sheet(SHEET_INGREDIENTES_MESTRE, MASTER_INGREDIENT_HEADERS, master_rows)

        if SHEET_INGREDIENTES_LEGACY in wb.sheetnames:
            del wb[SHEET_INGREDIENTES_LEGACY]
        if SHEET_CATALOGO_LEGACY in wb.sheetnames:
            del wb[SHEET_CATALOGO_LEGACY]

        if SHEET_INGREDIENTES_FICHA in wb.sheetnames:
            del wb[SHEET_INGREDIENTES_FICHA]
        ws_nf = wb.create_sheet(SHEET_INGREDIENTES_FICHA)
        _write_sheet(ws_nf, RECIPE_LINE_HEADERS, new_recipe)

        if SHEET_PRODUTOS in wb.sheetnames:
            _write_sheet(wb[SHEET_PRODUTOS], PRODUCT_HEADERS, products)

        fichas_db.save_workbook_safe(wb)
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        raise
    return warnings
